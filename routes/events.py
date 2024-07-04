from fastapi import APIRouter, HTTPException, status,Depends
from typing import Any, List, Union
from fastapi.responses import FileResponse
from typing_extensions import Annotated
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
import datetime
from openpyxl import Workbook
from models import (Events,User,EventsStudents)
from schemas import (
    events as events_schema
)
from modules import (
    auth as auth_module
)
from pony.orm import db_session, commit

router = APIRouter(
    prefix="/events",
    responses={404: {"description": "Not found"}},
)

@router.post("/create",status_code=status.HTTP_201_CREATED)
@db_session
def create_event(event:events_schema.event_in,current_user: Annotated[str, Depends(auth_module.get_current_user_id)]) :
    event =  Events(
        name=event.name,
        user=current_user
    )

    return {
        "event": event.name,
    }

@router.get("",status_code=status.HTTP_200_OK)
@db_session
def get_events(current_user: Annotated[str, Depends(auth_module.get_current_user_id)]):
    user = User.get(ID=current_user)
    data = []
    careerID = user.careerID

    # users_by_career_id = User.select(lambda u: u.careerID == careerID)
    ##get te events created by the user
    for event in user.events:
        data.append({
            "ID": event.ID,
            "name": event.name,
            "user": event.user.firstName + " " + event.user.middleName + " " + event.user.lastName,
            "createdAt": datetime.datetime.strftime(event.createdAt, "%d/%m/%Y")
        })

    # for user in users_by_career_id:
    #     for event in user.events:
    #         data.append({
    #             "ID": event.ID,
    #             "name": event.name,
    #             "user": event.user.firstName + " " + event.user.middleName + " " + event.user.lastName,
    #             "createdAt": datetime.datetime.strftime(event.createdAt, "%d/%m/%Y")
    #         })

    return data

@router.get("/{id}",status_code=status.HTTP_200_OK)
@db_session
def get_events(id:int):
    event = Events.get(ID=id)

    students_assist = []
    for event_student in event.events_students:
        students_assist.append({
            "ID": event_student.ID,
            "student": event_student.student.firstName + " " + event_student.student.middleName + " " + event_student.student.lastName,
            "control_number": event_student.student.controlNumber,
            "quantity_assist": event_student.quantity_assist,
            "user":event_student.user_id.firstName + " " + event_student.user_id.middleName + " " + event_student.user_id.lastName,
            "createdAt": datetime.datetime.strftime(event_student.createdAt, "%d/%m/%Y %H:%M:%S"),
            "career": event_student.student.careerID.name
        })

    return {
        "event":{
            "name": event.name,
            "user": event.user.firstName + " " + event.user.middleName + " " + event.user.lastName,
            "createdAt": datetime.datetime.strftime(event.createdAt, "%d/%m/%Y")
        },
        "students_assist": students_assist
    }

@router.get("/assist/{id}",status_code=status.HTTP_200_OK)
@db_session
def get_students_assist(id:int):
    event_student = EventsStudents.get(ID=id)
    return {
        "student": event_student.student.firstName + " " + event_student.student.middleName + " " + event_student.student.lastName,
        "control_number": event_student.student.controlNumber,
        "quantity_assist": event_student.quantity_assist,
        "career": event_student.student.careerID.name,
        "user":event_student.user_id.firstName + " " + event_student.user_id.middleName + " " + event_student.user_id.lastName,
    }

@router.delete("/{id}")
@db_session
def delete_event(id:int):
    event = Events.get(ID=id)
    event.delete()
    commit()
    return {
        "message": "Event deleted"
    }


@router.put("/{id}",status_code=status.HTTP_200_OK)
@db_session
def update_event(id:int,event_to_update:events_schema.event_in):
    event = Events.get(ID=id)
    event.name = event_to_update.name
    commit()
    return {
        "message": "Event updated"
    }

@router.get("/detail/{id}",status_code=status.HTTP_200_OK)
@db_session
def get_event_detail(id:int):
    event = Events.get(ID=id)
    return {
        "event": event.name
    }

@router.put("/assist/{id}/{quantity_assist_update}",status_code=status.HTTP_200_OK)
@db_session
def update_assist(id:int, quantity_assist_update:int):
    event_student = EventsStudents.get(ID=id)
    print(event_student)
    event_student.quantity_assist = quantity_assist_update
    commit()

    return {
        "message": "Assist updated"
    }

@router.delete("/assist/{id}")
@db_session
def delete_assist(id:int):
    event_student = EventsStudents.get(ID=id)
    event_student.delete()
    commit()
    return {
        "message": "Assist deleted"
    }

@router.get("/download_report/{id}/{quantity}",status_code=status.HTTP_200_OK)
@db_session
def download_report(id:int,quantity:int):
    event = Events.get(ID=id)
    wb = Workbook()
    ws = wb.active
    ws.title = f"Lista de asistencia {event.name}"
    ws['A1'] = "Nombre"
    ws['B1'] = "No. de control"
    ws['C1'] = "Carrera"
    ws['D1'] = "Fecha de asistencia"
    ws['E1'] = "Cantidad de asistencias"
    ws['F1'] = "Usuario quien registro"

    row = 2
    for event_student in event.events_students:
       if event_student.quantity_assist >= quantity:
           ws[f"A{row}"] = event_student.student.firstName + " " + event_student.student.middleName + " " + event_student.student.lastName
           ws[f"B{row}"] = event_student.student.controlNumber
           ws[f"C{row}"] = event_student.student.careerID.name
           ws[f"D{row}"] = datetime.datetime.strftime(event_student.createdAt, "%d/%m/%Y %H:%M:%S")
           ws[f"E{row}"] = event_student.quantity_assist
           ws[f"F{row}"] = event_student.user_id.firstName + " " + event_student.user_id.middleName + " " + event_student.user_id.lastName
           row += 1

    wb.save("tmp/report.xlsx")
    return FileResponse("tmp/report.xlsx", filename=f"Lista de asistencia {event.name}.xlsx")